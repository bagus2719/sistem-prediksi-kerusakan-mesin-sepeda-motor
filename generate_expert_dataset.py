"""
Generator Dataset Pakar Motor Matic - MotorCare C4.5 (25 Gejala)
================================================================
Cakupan: Mesin (Mekanikal, Bahan Bakar, Pengapian, Pendinginan, Pelumasan) + CVT
Output : CSV + Excel (.xlsx) dengan sheet referensi rules
"""
import csv
import random

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False
    print("⚠️  openpyxl belum terinstall. Jalankan: pip install openpyxl")
    print("   File Excel tidak akan dibuat, hanya CSV.\n")

random.seed(42)

# === 25 GEJALA ===
GEJALA_LIST = [
    ("G01", "Motor sulit dihidupkan (starter & kick)"),
    ("G02", "Mesin mati mendadak saat berjalan"),
    ("G03", "Suara kasar tek-tek dari area head mesin"),
    ("G04", "Asap putih tebal dari knalpot"),
    ("G05", "Tenaga mesin terasa lemas / ngempos"),
    ("G06", "Mesin cepat panas / indikator suhu menyala"),
    ("G07", "Tarikan awal berat dan bergetar (gredek)"),
    ("G08", "Suara krecek dari area timing chain"),
    ("G09", "Lampu MIL / Check Engine menyala"),
    ("G10", "Konsumsi BBM sangat boros"),
    ("G11", "Mesin mbrebet / tersendat saat digas"),
    ("G12", "Suara berdecit dari area CVT"),
    ("G13", "Suara kasar klotok dari area CVT"),
    ("G14", "Cairan radiator cepat habis"),
    ("G15", "Oli mesin cepat berkurang tanpa rembesan"),
    ("G16", "Kick starter terasa ringan (tidak ada kompresi)"),
    ("G17", "Mesin mati setelah dipakai lama (panas)"),
    ("G18", "Kick starter macet tidak bisa diinjak"),
    ("G19", "Putaran gas tidak kembali / nyangkut"),
    ("G20", "Bau sangit / terbakar dari area mesin"),
    ("G21", "Rembesan oli dari bodi mesin"),
    ("G22", "Lampu indikator oli menyala"),
    ("G23", "Motor hidup tapi tidak bergerak maju"),
    ("G24", "RPM langsam naik turun sendiri (hunting)"),
    ("G25", "Getaran berlebih pada mesin saat idle"),
    ("G26", "Tidak ada percikan api di busi sama sekali"),
]
GEJALA_HEADERS = [g[0] for g in GEJALA_LIST]

# === DATABASE MOTOR (Data Real Dunia Nyata) ===
MOTORS = [
    # ── HONDA INJEKSI ──
    (1,  "HONDA", "Beat FI", "Injeksi"),
    (2,  "HONDA", "Beat Street", "Injeksi"),
    (3,  "HONDA", "Vario 125", "Injeksi"),
    (4,  "HONDA", "Vario 150", "Injeksi"),
    (5,  "HONDA", "Scoopy FI", "Injeksi"),
    (6,  "HONDA", "Genio", "Injeksi"),
    (7,  "HONDA", "PCX 160", "Injeksi"),
    (8,  "HONDA", "ADV 160", "Injeksi"),

    # ── HONDA KARBURATOR ──
    (9,  "HONDA", "Beat Karbu", "Karburator"),
    (10, "HONDA", "Vario 110 Karbu", "Karburator"),
    (11, "HONDA", "Scoopy Karbu", "Karburator"),

    # ── YAMAHA INJEKSI ──
    (12, "YAMAHA", "Mio M3", "Injeksi"),
    (13, "YAMAHA", "Mio S", "Injeksi"),
    (14, "YAMAHA", "Soul GT FI", "Injeksi"),
    (15, "YAMAHA", "NMAX 155", "Injeksi"),
    (16, "YAMAHA", "Aerox 155", "Injeksi"),
    (17, "YAMAHA", "Lexi", "Injeksi"),
    (18, "YAMAHA", "Freego", "Injeksi"),
    (19, "YAMAHA", "Fino FI", "Injeksi"),

    # ── YAMAHA KARBURATOR ──
    (20, "YAMAHA", "Mio Sporty", "Karburator"),
    (21, "YAMAHA", "Mio Soul Karbu", "Karburator"),
    (22, "YAMAHA", "Fino Karbu", "Karburator"),
    (23, "YAMAHA", "Xeon Karbu", "Karburator"),
]

# === ATURAN PAKAR ===
# Prinsip: Setiap gejala UTAMA hanya menunjuk ke 1 kerusakan per cabang tipe sistem.
# "utama" = selalu ada. "pendukung" = kadang ada (variasi).
# Pendukung TIDAK BOLEH menjadi utama penyakit lain di cabang yang sama.
#
# PEMETAAN (tidak boleh tabrakan):
# ┌───────────────────────────────────────────────────────┐
# │ KEDUANYA:                                             │
# │  G03→K15  G04→K10  G06→K03  G07→K09  G08→K16        │
# │  G12→K31  G13→K33  G14→K24  G17→K18  G18→K36        │
# │  G21→K35  G22→K30  G23→K37  G25→K38                 │
# ├───────────────────────────────────────────────────────┤
# │ INJEKSI:  G01→K06  G02→K05  G09→K08  G19→K20        │
# │           G24→K42  G26→K40(ECU Rusak)                 │
# ├───────────────────────────────────────────────────────┤
# │ KARBURATOR: G01→K04  G02→K02  G10→K01  G19→K44      │
# │             G24→K45  G26→K41(CDI Rusak)               │
# └───────────────────────────────────────────────────────┘
# Support-only (tidak pernah jadi utama): G05, G10(inj), G11, G15, G16, G20

EXPERT_RULES = [
    # ── KEDUANYA ─────────────────────────────────────────
    # Mekanikal Atas
    {"kode": "K15", "nama": "Bearing Noken As Aus", "tipe": "Keduanya",
     "kategori": "Mekanikal",
     "utama": ["G03"], "pendukung": ["G25"],
     "solusi": "Ganti bearing noken as dan periksa keausan camshaft."},

    {"kode": "K10", "nama": "Piston / Ring Piston Aus", "tipe": "Keduanya",
     "kategori": "Mekanikal",
     "utama": ["G04"], "pendukung": ["G15", "G05"],
     "solusi": "Korter blok silinder, ganti piston dan ring piston set."},

    {"kode": "K16", "nama": "Rantai Keteng Kendur", "tipe": "Keduanya",
     "kategori": "Mekanikal",
     "utama": ["G08"], "pendukung": [],
     "solusi": "Ganti tensioner dan rantai keteng (timing chain)."},

    {"kode": "K36", "nama": "Piston Macet (Seizure)", "tipe": "Keduanya",
     "kategori": "Mekanikal",
     "utama": ["G18"], "pendukung": ["G20"],
     "solusi": "Bongkar mesin, korter ulang silinder, ganti piston baru."},

    {"kode": "K38", "nama": "Kruk As / Bearing Besar Oblak", "tipe": "Keduanya",
     "kategori": "Mekanikal",
     "utama": ["G25"], "pendukung": ["G20"],
     "solusi": "Ganti bearing besar kruk as (crankshaft bearing)."},

    {"kode": "K13", "nama": "Klep (Valve) Bocor", "tipe": "Keduanya",
     "kategori": "Mekanikal",
     "utama": ["G16"], "pendukung": ["G05"],
     "solusi": "Skir klep, ganti seal klep, dan setel celah klep."},

    # Pengapian
    {"kode": "K18", "nama": "Koil Pengapian Rusak", "tipe": "Keduanya",
     "kategori": "Pengapian",
     "utama": ["G17"], "pendukung": ["G11"],
     "solusi": "Ganti koil pengapian dengan yang baru sesuai spesifikasi."},

    # Pendinginan
    {"kode": "K03", "nama": "Overheat (Radiator Tersumbat)", "tipe": "Keduanya",
     "kategori": "Pendinginan",
     "utama": ["G06"], "pendukung": ["G14"],
     "solusi": "Flush radiator, ganti coolant, periksa water pump."},

    {"kode": "K24", "nama": "Kipas Radiator Mati", "tipe": "Keduanya",
     "kategori": "Pendinginan",
     "utama": ["G14"], "pendukung": [],
     "solusi": "Ganti motor kipas radiator atau periksa relay/fuse kipas."},

    # Pelumasan
    {"kode": "K30", "nama": "Pompa Oli Mesin Lemah", "tipe": "Keduanya",
     "kategori": "Pelumasan",
     "utama": ["G22"], "pendukung": ["G20"],
     "solusi": "Ganti pompa oli mesin dan periksa saluran oli."},

    {"kode": "K35", "nama": "Seal / Paking Mesin Bocor", "tipe": "Keduanya",
     "kategori": "Pelumasan",
     "utama": ["G21"], "pendukung": ["G15"],
     "solusi": "Ganti gasket/paking head dan seal mesin yang bocor."},

    # CVT
    {"kode": "K09", "nama": "Roller CVT Aus", "tipe": "Keduanya",
     "kategori": "CVT",
     "utama": ["G07"], "pendukung": ["G13"],
     "solusi": "Ganti roller CVT satu set sesuai berat standar."},

    {"kode": "K31", "nama": "V-Belt CVT Retak / Putus", "tipe": "Keduanya",
     "kategori": "CVT",
     "utama": ["G12"], "pendukung": [],
     "solusi": "Ganti V-Belt CVT dengan ukuran standar."},

    {"kode": "K33", "nama": "Kampas Ganda Aus", "tipe": "Keduanya",
     "kategori": "CVT",
     "utama": ["G13"], "pendukung": [],
     "solusi": "Ganti kampas ganda (clutch shoe) CVT satu set."},

    {"kode": "K37", "nama": "Per CVT Lemah / Patah", "tipe": "Keduanya",
     "kategori": "CVT",
     "utama": ["G23"], "pendukung": [],
     "solusi": "Ganti per CVT (center spring) dengan yang baru."},

    # ── ENTRI ALTERNATIF (Gejala → Penyakit dari sudut pandang berbeda) ──
    # Tujuan: Agar setiap gejala yang bisa dipilih user PASTI punya jalur
    # yang benar di pohon C4.5, meskipun user hanya memilih 1 gejala.
    # Kode penyakit SAMA, tapi gejala utama BERBEDA.

    # G15 (Oli berkurang tanpa rembesan) → Piston aus (oli terbakar internal)
    {"kode": "K10", "nama": "Piston / Ring Piston Aus", "tipe": "Keduanya",
     "kategori": "Mekanikal",
     "utama": ["G15"], "pendukung": ["G04"],
     "solusi": "Korter blok silinder, ganti piston dan ring piston set."},

    # G05 (Tenaga lemas/ngempos) → Klep bocor (kompresi hilang)
    {"kode": "K13", "nama": "Klep (Valve) Bocor", "tipe": "Keduanya",
     "kategori": "Mekanikal",
     "utama": ["G05"], "pendukung": ["G16"],
     "solusi": "Skir klep, ganti seal klep, dan setel celah klep."},

    # G11 (Mbrebet/tersendat) → Koil pengapian rusak (api intermittent)
    {"kode": "K18", "nama": "Koil Pengapian Rusak", "tipe": "Keduanya",
     "kategori": "Pengapian",
     "utama": ["G11"], "pendukung": [],
     "solusi": "Ganti koil pengapian dengan yang baru sesuai spesifikasi."},

    # G20 (Bau sangit/terbakar) → Piston macet (gesekan berlebih)
    {"kode": "K36", "nama": "Piston Macet (Seizure)", "tipe": "Keduanya",
     "kategori": "Mekanikal",
     "utama": ["G20"], "pendukung": ["G18"],
     "solusi": "Bongkar mesin, korter ulang silinder, ganti piston baru."},

    # ── INJEKSI ──────────────────────────────────────────
    {"kode": "K06", "nama": "Injektor Kotor / Tersumbat", "tipe": "Injeksi",
     "kategori": "Bahan Bakar",
     "utama": ["G01"], "pendukung": ["G10", "G11"],
     "solusi": "Bersihkan injektor dengan injector cleaner atau ganti baru."},

    # G10 (BBM boros di Injeksi) → Injektor kotor (semprotan tidak optimal)
    {"kode": "K06", "nama": "Injektor Kotor / Tersumbat", "tipe": "Injeksi",
     "kategori": "Bahan Bakar",
     "utama": ["G10"], "pendukung": ["G01"],
     "solusi": "Bersihkan injektor dengan injector cleaner atau ganti baru."},

    {"kode": "K05", "nama": "Fuel Pump Lemah", "tipe": "Injeksi",
     "kategori": "Bahan Bakar",
     "utama": ["G02"], "pendukung": [],
     "solusi": "Periksa tekanan fuel pump, ganti rotak atau assy."},

    {"kode": "K08", "nama": "Sensor TPS / ECU Error", "tipe": "Injeksi",
     "kategori": "Bahan Bakar",
     "utama": ["G09"], "pendukung": ["G11"],
     "solusi": "Reset ECU, kalibrasi TPS, atau ganti sensor TPS."},

    {"kode": "K20", "nama": "Throttle Body Kotor", "tipe": "Injeksi",
     "kategori": "Bahan Bakar",
     "utama": ["G19"], "pendukung": ["G24"],
     "solusi": "Bersihkan throttle body dan ISC/IACV dengan carb cleaner."},

    {"kode": "K42", "nama": "ISC / IACV Kotor", "tipe": "Injeksi",
     "kategori": "Bahan Bakar",
     "utama": ["G24"], "pendukung": [],
     "solusi": "Bersihkan atau ganti ISC/IACV (Idle Speed Control)."},

    {"kode": "K40", "nama": "ECU Rusak", "tipe": "Injeksi",
     "kategori": "Pengapian",
     "utama": ["G26"], "pendukung": ["G09"],
     "solusi": "Periksa ECU dengan scanner, reset atau ganti ECU."},

    # ── KARBURATOR ───────────────────────────────────────
    {"kode": "K04", "nama": "Karburator Kotor", "tipe": "Karburator",
     "kategori": "Bahan Bakar",
     "utama": ["G01"], "pendukung": ["G10", "G11"],
     "solusi": "Bongkar karburator, bersihkan pilot jet dan main jet."},

    {"kode": "K02", "nama": "Aki Soak / Lemah", "tipe": "Karburator",
     "kategori": "Kelistrikan",
     "utama": ["G02"], "pendukung": [],
     "solusi": "Cas ulang aki atau ganti aki baru, periksa kiprok/regulator."},

    {"kode": "K01", "nama": "Busi Bermasalah", "tipe": "Karburator",
     "kategori": "Pengapian",
     "utama": ["G10"], "pendukung": ["G11", "G05"],
     "solusi": "Ganti busi baru sesuai spesifikasi, cek gap/celah busi."},

    {"kode": "K44", "nama": "Kabel Gas Aus / Macet", "tipe": "Karburator",
     "kategori": "Bahan Bakar",
     "utama": ["G19"], "pendukung": [],
     "solusi": "Lumasi atau ganti kabel gas (throttle cable)."},

    {"kode": "K45", "nama": "Setelan Angin Karburator Salah", "tipe": "Karburator",
     "kategori": "Bahan Bakar",
     "utama": ["G24"], "pendukung": ["G11"],
     "solusi": "Setel ulang setelan angin (pilot screw) karburator."},

    {"kode": "K41", "nama": "CDI Rusak", "tipe": "Karburator",
     "kategori": "Pengapian",
     "utama": ["G26"], "pendukung": [],
     "solusi": "Ganti CDI (Capacitor Discharge Ignition) dengan yang baru."},
]


def normalize_kode(kode):
    return kode.split("_")[0]


def generate_rows_for_rule(rule, id_start):
    """Generate variasi baris: utama saja, utama+sebagian, utama+semua pendukung."""
    rows = []
    valid_motors = [m for m in MOTORS if rule["tipe"] == "Keduanya" or m[3] == rule["tipe"]]
    kode_out = normalize_kode(rule["kode"])

    patterns = []
    # Pola 1: Hanya gejala utama
    patterns.append(list(rule["utama"]))
    # Pola 2: Utama + semua pendukung
    if rule["pendukung"]:
        patterns.append(rule["utama"] + rule["pendukung"])
        # Pola 3: Utama + tiap pendukung individual
        for p in rule["pendukung"]:
            pat = rule["utama"] + [p]
            if pat not in patterns:
                patterns.append(pat)

    id_counter = id_start
    for pattern in patterns:
        sample_motors = random.sample(valid_motors, min(len(valid_motors), 3))
        for motor in sample_motors:
            gejala_data = {g: 0 for g in GEJALA_HEADERS}
            for g in pattern:
                if g in gejala_data:
                    gejala_data[g] = 1

            row = [id_counter, motor[0], motor[1], motor[2], motor[3]]
            for g in GEJALA_HEADERS:
                row.append(gejala_data[g])
            row.extend([kode_out, rule["nama"], rule["solusi"], ""])
            rows.append(row)
            id_counter += 1
    return rows


def validate(dataset, headers):
    """Cek kontradiksi: pola gejala+tipe sama tapi kerusakan beda."""
    seen = {}
    conflicts = 0
    tipe_idx = headers.index("Tipe_Sistem")
    kode_idx = headers.index("Kode_Kerusakan")
    g_start = headers.index("G01")
    g_end = headers.index("G26") + 1
    for row in dataset:
        key = f"{row[tipe_idx]}|{'|'.join(str(row[i]) for i in range(g_start, g_end))}"
        kode = row[kode_idx]
        if key in seen and seen[key] != kode:
            print(f"  ❌ KONFLIK: {kode} vs {seen[key]}")
            conflicts += 1
        else:
            seen[key] = kode
    return conflicts


def write_excel(final_dataset, headers):
    """Buat file Excel dengan 2 sheet: Dataset dan Ref Rules."""
    if not HAS_OPENPYXL:
        return

    wb = Workbook()

    # --- Sheet 1: Dataset ---
    ws1 = wb.active
    ws1.title = "Dataset"
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'))

    for col_idx, h in enumerate(headers, 1):
        cell = ws1.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    for row_idx, row_data in enumerate(final_dataset, 2):
        for col_idx, val in enumerate(row_data, 1):
            cell = ws1.cell(row=row_idx, column=col_idx, value=val)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center")

    # Auto-width
    for col in ws1.columns:
        max_len = max(len(str(c.value or "")) for c in col)
        ws1.column_dimensions[col[0].column_letter].width = min(max_len + 3, 40)

    # --- Sheet 2: Ref Gejala ---
    ws2 = wb.create_sheet("Ref Gejala")
    gejala_headers = ["Kode", "Nama Gejala"]
    green_fill = PatternFill(start_color="059669", end_color="059669", fill_type="solid")
    for col_idx, h in enumerate(gejala_headers, 1):
        cell = ws2.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = green_fill
        cell.border = thin_border
    for row_idx, (kode, nama) in enumerate(GEJALA_LIST, 2):
        ws2.cell(row=row_idx, column=1, value=kode).border = thin_border
        ws2.cell(row=row_idx, column=2, value=nama).border = thin_border
    ws2.column_dimensions['A'].width = 8
    ws2.column_dimensions['B'].width = 55

    # --- Sheet 3: Ref Rules ---
    ws3 = wb.create_sheet("Ref Rules")
    rule_headers = ["Kode", "Nama Kerusakan", "Tipe Sistem", "Kategori",
                     "Gejala Utama", "Gejala Pendukung", "Solusi"]
    amber_fill = PatternFill(start_color="D97706", end_color="D97706", fill_type="solid")
    for col_idx, h in enumerate(rule_headers, 1):
        cell = ws3.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = amber_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    for row_idx, rule in enumerate(EXPERT_RULES, 2):
        kode = normalize_kode(rule["kode"])
        ws3.cell(row=row_idx, column=1, value=kode).border = thin_border
        ws3.cell(row=row_idx, column=2, value=rule["nama"]).border = thin_border
        ws3.cell(row=row_idx, column=3, value=rule["tipe"]).border = thin_border
        ws3.cell(row=row_idx, column=4, value=rule["kategori"]).border = thin_border
        ws3.cell(row=row_idx, column=5, value=", ".join(rule["utama"])).border = thin_border
        ws3.cell(row=row_idx, column=6, value=", ".join(rule["pendukung"]) or "-").border = thin_border
        ws3.cell(row=row_idx, column=7, value=rule["solusi"]).border = thin_border

    for col_letter, width in [('A',8),('B',35),('C',14),('D',14),('E',16),('F',20),('G',60)]:
        ws3.column_dimensions[col_letter].width = width

    # --- Sheet 4: Pemetaan Gejala-Kerusakan ---
    ws4 = wb.create_sheet("Pemetaan")
    rose_fill = PatternFill(start_color="E11D48", end_color="E11D48", fill_type="solid")
    map_headers = ["Gejala", "Tipe Sistem", "Kerusakan (Kode)", "Kerusakan (Nama)", "Peran"]
    for col_idx, h in enumerate(map_headers, 1):
        cell = ws4.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = rose_fill
        cell.border = thin_border

    row_idx = 2
    for rule in EXPERT_RULES:
        kode = normalize_kode(rule["kode"])
        tipe = rule["tipe"]
        for g in rule["utama"]:
            ws4.cell(row=row_idx, column=1, value=g).border = thin_border
            ws4.cell(row=row_idx, column=2, value=tipe).border = thin_border
            ws4.cell(row=row_idx, column=3, value=kode).border = thin_border
            ws4.cell(row=row_idx, column=4, value=rule["nama"]).border = thin_border
            ws4.cell(row=row_idx, column=5, value="UTAMA").border = thin_border
            row_idx += 1
        for g in rule["pendukung"]:
            ws4.cell(row=row_idx, column=1, value=g).border = thin_border
            ws4.cell(row=row_idx, column=2, value=tipe).border = thin_border
            ws4.cell(row=row_idx, column=3, value=kode).border = thin_border
            ws4.cell(row=row_idx, column=4, value=rule["nama"]).border = thin_border
            ws4.cell(row=row_idx, column=5, value="Pendukung").border = thin_border
            row_idx += 1

    for col_letter, width in [('A',8),('B',14),('C',16),('D',35),('E',12)]:
        ws4.column_dimensions[col_letter].width = width

    excel_path = "dataset_motor_pakar_murni.xlsx"
    wb.save(excel_path)
    print(f"📊 File Excel '{excel_path}' berhasil dibuat!")
    print(f"   Sheet: Dataset | Ref Gejala | Ref Rules | Pemetaan")


def main():
    OUTPUT_CSV = 'dataset_motor_pakar_murni.csv'
    headers = (["ID", "ID_Motor", "Merk", "Model", "Tipe_Sistem"] +
               GEJALA_HEADERS +
               ["Kode_Kerusakan", "Nama_Kerusakan", "Solusi", "Inferensi_FC"])

    # Generate basis dari semua rules
    all_rows = []
    id_counter = 1
    for rule in EXPERT_RULES:
        new = generate_rows_for_rule(rule, id_counter)
        all_rows.extend(new)
        id_counter += len(new)

    print(f"Baris basis: {len(all_rows)}")

    # Perbanyak hingga ~500
    target = 500
    final = list(all_rows)
    while len(final) < target:
        rule = random.choice(EXPERT_RULES)
        new = generate_rows_for_rule(rule, len(final) + 1)
        for r in new:
            if len(final) >= target:
                break
            r[0] = len(final) + 1
            final.append(r)

    random.shuffle(final)
    for i, row in enumerate(final):
        row[0] = i + 1

    # Validasi
    print(f"Total baris: {len(final)}")
    print("Memeriksa kontradiksi...")
    c = validate(final, headers)
    print("✅ LULUS!" if c == 0 else f"⚠️ {c} kontradiksi!")

    # Distribusi
    kode_idx = headers.index("Kode_Kerusakan")
    dist = {}
    for row in final:
        k = row[kode_idx]
        dist[k] = dist.get(k, 0) + 1
    print("\nDistribusi kerusakan:")
    for k in sorted(dist.keys()):
        print(f"  {k}: {dist[k]} baris")

    # Tulis CSV
    with open(OUTPUT_CSV, 'w', newline='', encoding='utf-8') as f:
        writer = csv.writer(f)
        writer.writerow(headers)
        writer.writerows(final)
    print(f"\n📄 File CSV '{OUTPUT_CSV}' berhasil dibuat!")

    # Tulis Excel
    write_excel(final, headers)

    print("\n🎉 SELESAI! Langkah selanjutnya:")
    print("  1. Truncate tabel trainings di database")
    print("  2. Import CSV via Admin > Data Training")
    print("  3. Update tabel gejalas (jalankan seeder baru untuk 25 gejala)")
    print("  4. Klik Generate Model C4.5")
    print("  5. Uji prediksi!")


if __name__ == "__main__":
    main()
